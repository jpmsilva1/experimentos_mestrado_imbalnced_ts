#!/usr/bin/env python3
"""
sync_from_excel.py — Sync paper registry from Fichamento_Artigos.xlsx to docs/paper_registry.md.

Reads the Excel file from Google Drive, extracts all paper metadata, and generates
a structured markdown document listing all 43 papers with their metadata.

Usage:
    python3 scripts/sync_from_excel.py

Requirements:
    pip install openpyxl
"""

import sys
from pathlib import Path
from datetime import date

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

REPO_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = Path(
    "/Users/joaopms/Library/CloudStorage/GoogleDrive-jpms5@cin.ufpe.br/"
    "My Drive/Papers Mestrado - João/Fichamento_Artigos.xlsx"
)
OUTPUT_PATH = REPO_ROOT / "docs" / "paper_registry.md"


def main():
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at: {EXCEL_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in ws[1]]

    papers = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = list(row)
        if all(v is None for v in row_data):
            continue
        paper = {}
        for i, h in enumerate(headers):
            if h and i < len(row_data):
                paper[str(h)] = row_data[i]
        paper["excel_row"] = row_idx
        papers.append(paper)

    # Build markdown
    lines = [
        "---",
        f"# Paper Registry",
        "",
        f"> Auto-generated from `Fichamento_Artigos.xlsx` on {date.today().isoformat()}.",
        f"> **Total papers**: {len(papers)}",
        "",
        "---",
        "",
    ]

    # Group: papers with code vs without
    with_code = [p for p in papers if p.get("Disponibilidade de Códigos?") == "Sim"]
    without_code = [p for p in papers if p.get("Disponibilidade de Códigos?") != "Sim"]

    lines.append(f"## Papers with Code Available ({len(with_code)})")
    lines.append("")
    lines.append("| # | Title | Year | Authors | Code Link |")
    lines.append("|---|-------|------|---------|-----------|")
    for i, p in enumerate(with_code, 1):
        title = str(p.get("Título", "N/A"))[:80]
        year = p.get("Ano", "N/A")
        authors = str(p.get("Autores", "N/A"))[:50]
        code = p.get("Link para códigos", "—")
        code_link = f"[link]({code})" if code and code != "—" else "—"
        lines.append(f"| {i} | {title} | {year} | {authors} | {code_link} |")

    lines.append("")
    lines.append(f"## Papers without Code ({len(without_code)})")
    lines.append("")
    lines.append("| # | Title | Year | Authors |")
    lines.append("|---|-------|------|---------|")
    for i, p in enumerate(without_code, 1):
        title = str(p.get("Título", "N/A"))[:80]
        year = p.get("Ano", "N/A")
        authors = str(p.get("Autores", "N/A"))[:50]
        lines.append(f"| {i} | {title} | {year} | {authors} |")

    lines.append("")

    # Write output
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text("\n".join(lines), encoding="utf-8")
    print(f"✅ Paper registry written to: {OUTPUT_PATH}")
    print(f"   {len(with_code)} papers with code, {len(without_code)} without code")


if __name__ == "__main__":
    main()
